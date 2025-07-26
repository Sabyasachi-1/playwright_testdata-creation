import openpyxl
import os

def write_confirmation(file_name, row_index, booking_id, status, date):
    base_path = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_path, "..", "testdata", file_name)

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    row_excel = row_index + 2
    sheet.cell(row=row_excel, column=14).value = booking_id
    sheet.cell(row=row_excel, column=15).value = status
    sheet.cell(row=row_excel, column=16).value = date
    wb.save(file_path)