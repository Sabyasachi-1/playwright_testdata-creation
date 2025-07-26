from openpyxl import load_workbook
import os
def read_ticket_data(file_name):
    base_path = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_path, "..", "testdata", file_name)

    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    data =[]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data